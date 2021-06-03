import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QTableWidgetItem, QVBoxLayout, QWidget, QPushButton, QFileDialog, QTableWidget,QHBoxLayout
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import QRect, pyqtSlot
import csv
# from PyQt5.sip import array
import openpyxl
from pathlib import Path

class App(QWidget):

    def __init__(self):
        super().__init__()
        self.title = 'Excel plugin'
        self.left = 10
        self.top = 10
        self.width = 1000
        self.height = 1000
        layout=QHBoxLayout()


        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        
        button = QPushButton('upload')
        button.setToolTip('This is an example button')
        button.move(100,70)
        button.clicked.connect(self.on_click)
        layout.addWidget(button)
        self.tableWidget= QTableWidget()
        self.tableWidget.setGeometry(QRect(310,10,311,161))
        # self.layout.addWidget(self.tableWidget)
        layout.addWidget(self.tableWidget)
        self.setLayout(layout)
        self.show()

    def openFileNameDialog(self):
        self.tableWidget.setRowCount(158)
        self.tableWidget.setColumnCount(12)
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","All Files (*);;Python Files (*.py)", options=options)
        # print(file)
        xlsx_file = Path('SimData', file)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        print(wb_obj)
        sheet= wb_obj.active
        print(sheet)
        rowCounter=0
        x=[[]]
        arr = [[cell.value for cell in row] for row in sheet.iter_rows(max_row=158)]
        rowCount=0
        for i in arr:
            colCount=0
            for j in i:
                self.tableWidget.setItem(rowCount,colCount,QtWidgets.QTableWidgetItem(arr[rowCount][colCount]))
                colCount+=1
            rowCount+=1

        # for row in sheet.iter_rows(max_row=158):
        #     colCounter=0
        #     for cell in row:
        #             if(cell.value != None):
        #                 self.tableWidget.setItem(rowCounter,colCounter,QtWidgets.QTableWidgetItem(cell.value))
        #                 print(QtWidgets.QTableWidgetItem(cell.value))
        #             colCounter+=1
        #             print(cell.value, end=" ")    
        #     rowCounter+=1
        #     print()

        # with open(file, newline='') as csvfile:
        #     csv_reader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        #     rowNum=0
        #     for row in csv_reader:
        #         # print(row)
        #         Col=0
        #         while Col<len(row):
        #             # item= QtWidgets.QTableWidgetItem(row[counter][i],0)
        #             self.tableWidget.setItem(rowNum,Col,QtWidgets.QTableWidgetItem(row[Col]))
        #             Col+=1
        #     rowNum+=1
        self.tableWidget.move(0,0)
        self.tableWidget.doubleClicked.connect(self.on_click)

    @pyqtSlot()
    def on_click(self):
        self.openFileNameDialog()
        print("\n")
        for currentQTableWidgetItem in self.tableWidget.selectedItems():
            print(currentQTableWidgetItem.row(), currentQTableWidgetItem.column(), currentQTableWidgetItem.text())

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())
